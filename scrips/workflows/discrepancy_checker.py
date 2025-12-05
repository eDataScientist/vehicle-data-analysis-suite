import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
import io
from typing import Dict, List, Tuple, Optional, Union


class DiscrepancyChecker:
    """Excel/CSV processor with conditional formatting for discrepancy checking"""

    def __init__(self):
        self.df = None
        self.sheet_name = None
        self.available_sheets = None
        # For two-file mode
        self.primary_df = None
        self.join_df = None
        self.primary_sheets = None
        self.join_sheets = None
        self.merged_df = None

    def load_data(self, file_data: Union[io.BytesIO, str], file_type: str = 'xlsx') -> List[str]:
        """
        Load data from uploaded file and return available sheets

        Args:
            file_data: File data (BytesIO for uploaded files)
            file_type: 'xlsx', 'xls', or 'csv'

        Returns:
            List of sheet names (for Excel) or ['Data'] for CSV
        """
        if file_type in ['xlsx', 'xls']:
            wb = openpyxl.load_workbook(file_data)
            self.available_sheets = wb.sheetnames
            return self.available_sheets
        else:
            # CSV file
            self.df = pd.read_csv(file_data)
            self.available_sheets = ['Data']
            self.sheet_name = 'Data'
            return ['Data']

    def select_sheet(self, sheet_name: str, file_data: Union[io.BytesIO, str], file_type: str = 'xlsx') -> pd.DataFrame:
        """
        Select and load a specific sheet

        Args:
            sheet_name: Name of the sheet to load
            file_data: File data
            file_type: File type

        Returns:
            DataFrame with the selected sheet data
        """
        if file_type in ['xlsx', 'xls']:
            self.df = pd.read_excel(file_data, sheet_name=sheet_name)
        else:
            # Already loaded for CSV
            pass

        self.sheet_name = sheet_name

        # Standardize VIN header if present
        if 'Vin' in self.df.columns:
            self.df.rename(columns={'Vin': 'VIN'}, inplace=True)

        return self.df

    def get_columns(self) -> List[str]:
        """Get list of available columns"""
        if self.df is not None:
            return list(self.df.columns)
        return []
    
    # Two-file mode methods
    def load_primary_data(self, file_data: Union[io.BytesIO, str], file_type: str = 'xlsx') -> List[str]:
        """
        Load primary data file and return available sheets
        
        Args:
            file_data: File data (BytesIO for uploaded files)
            file_type: 'xlsx', 'xls', or 'csv'
            
        Returns:
            List of sheet names (for Excel) or ['Data'] for CSV
        """
        if file_type in ['xlsx', 'xls']:
            wb = openpyxl.load_workbook(file_data)
            self.primary_sheets = wb.sheetnames
            return self.primary_sheets
        else:
            # CSV file
            self.primary_df = pd.read_csv(file_data)
            self.primary_sheets = ['Data']
            return ['Data']
    
    def load_join_data(self, file_data: Union[io.BytesIO, str], file_type: str = 'xlsx') -> List[str]:
        """
        Load join data file and return available sheets
        
        Args:
            file_data: File data (BytesIO for uploaded files)
            file_type: 'xlsx', 'xls', or 'csv'
            
        Returns:
            List of sheet names (for Excel) or ['Data'] for CSV
        """
        if file_type in ['xlsx', 'xls']:
            wb = openpyxl.load_workbook(file_data)
            self.join_sheets = wb.sheetnames
            return self.join_sheets
        else:
            # CSV file
            self.join_df = pd.read_csv(file_data)
            self.join_sheets = ['Data']
            return ['Data']
    
    def select_primary_sheet(self, sheet_name: str, file_data: Union[io.BytesIO, str], file_type: str = 'xlsx') -> pd.DataFrame:
        """
        Select and load a specific sheet from primary data
        
        Args:
            sheet_name: Name of the sheet to load
            file_data: File data
            file_type: File type
            
        Returns:
            DataFrame with the selected sheet data
        """
        if file_type in ['xlsx', 'xls']:
            self.primary_df = pd.read_excel(file_data, sheet_name=sheet_name)
        else:
            # Already loaded for CSV
            pass
        
        # Standardize chassis/VIN headers if present
        if 'Vin' in self.primary_df.columns:
            self.primary_df.rename(columns={'Vin': 'VIN'}, inplace=True)
        if 'Chassis' in self.primary_df.columns:
            self.primary_df.rename(columns={'Chassis': 'chassis_no'}, inplace=True)
        if 'chassis_no' not in self.primary_df.columns and 'VIN' in self.primary_df.columns:
            self.primary_df.rename(columns={'VIN': 'chassis_no'}, inplace=True)
            
        return self.primary_df
    
    def select_join_sheet(self, sheet_name: str, file_data: Union[io.BytesIO, str], file_type: str = 'xlsx') -> pd.DataFrame:
        """
        Select and load a specific sheet from join data
        
        Args:
            sheet_name: Name of the sheet to load
            file_data: File data
            file_type: File type
            
        Returns:
            DataFrame with the selected sheet data
        """
        if file_type in ['xlsx', 'xls']:
            self.join_df = pd.read_excel(file_data, sheet_name=sheet_name)
        else:
            # Already loaded for CSV
            pass
        
        # Standardize chassis/VIN headers if present
        if 'Vin' in self.join_df.columns:
            self.join_df.rename(columns={'Vin': 'VIN'}, inplace=True)
        if 'Chassis' in self.join_df.columns:
            self.join_df.rename(columns={'Chassis': 'chassis_no'}, inplace=True)
        if 'chassis_no' not in self.join_df.columns and 'VIN' in self.join_df.columns:
            self.join_df.rename(columns={'VIN': 'chassis_no'}, inplace=True)
            
        return self.join_df
    
    def get_primary_columns(self) -> List[str]:
        """Get list of available columns from primary data"""
        if self.primary_df is not None:
            return list(self.primary_df.columns)
        return []
    
    def get_join_columns(self) -> List[str]:
        """Get list of available columns from join data"""
        if self.join_df is not None:
            return list(self.join_df.columns)
        return []
    
    def join_data(self, primary_chassis_col: str, join_chassis_col: str, join_type: str = 'left') -> pd.DataFrame:
        """
        Join primary and join data based on chassis numbers
        
        Args:
            primary_chassis_col: Chassis column name in primary data
            join_chassis_col: Chassis column name in join data
            join_type: Type of join ('left', 'right', 'inner', 'outer')
            
        Returns:
            Merged DataFrame
        """
        if self.primary_df is None or self.join_df is None:
            raise ValueError("Both primary and join data must be loaded")
        
        # Ensure chassis columns exist
        if primary_chassis_col not in self.primary_df.columns:
            raise ValueError(f"Primary chassis column '{primary_chassis_col}' not found in primary data. Available: {list(self.primary_df.columns)}")
        if join_chassis_col not in self.join_df.columns:
            raise ValueError(f"Join chassis column '{join_chassis_col}' not found in join data. Available: {list(self.join_df.columns)}")
        
        # Perform the join
        self.merged_df = self.primary_df.merge(
            self.join_df,
            left_on=primary_chassis_col,
            right_on=join_chassis_col,
            how=join_type,
            suffixes=('_primary', '_join')
        )
        
        # Set merged_df as the working dataframe for processing
        self.df = self.merged_df.copy()
        self.sheet_name = "Merged_Data"
        
        return self.merged_df
    
    def get_merged_columns(self) -> List[str]:
        """Get list of available columns from merged data"""
        if self.merged_df is not None:
            return list(self.merged_df.columns)
        return []

    def process_comparison_mode(self, col1: str, col2: str, threshold: float, highlight_mode: str = "entire_row", dividend_col: str = None) -> io.BytesIO:
        """
        Process comparison mode - compare two columns within threshold

        Args:
            col1: First column name
            col2: Second column name  
            threshold: Threshold percentage
            highlight_mode: "entire_row" or "percentage_column"
            dividend_col: Column to use as dividend (denominator). If None, defaults to col2

        Returns:
            BytesIO object containing the Excel file
        """
        if self.df is None:
            raise ValueError("No data loaded")

        # Default dividend_col to col2 if not specified (maintains backward compatibility)
        if dividend_col is None:
            dividend_col = col2
        
        # Validate dividend_col
        if dividend_col not in [col1, col2]:
            raise ValueError(f"dividend_col must be either '{col1}' or '{col2}', got '{dividend_col}'")

        # Add placeholder column for percentage difference - will be replaced with formulas
        self.df['Percentage_Difference'] = 0

        return self._create_excel_output(mode=1, col1=col1, col2=col2, threshold=threshold, highlight_mode=highlight_mode, dividend_col=dividend_col)

    def process_range_mode(self, anchor_col: str, low_col: str, high_col: str, highlight_mode: str = "entire_row") -> io.BytesIO:
        """
        Process range mode - check if anchor value is between low and high

        Args:
            anchor_col: Anchor column name
            low_col: Low boundary column name
            high_col: High boundary column name
            highlight_mode: "entire_row" or "anchor_column"

        Returns:
            BytesIO object containing the Excel file
        """
        if self.df is None:
            raise ValueError("No data loaded")

        return self._create_excel_output(mode=2, anchor_col=anchor_col, low_col=low_col, high_col=high_col, highlight_mode=highlight_mode)

    def process_absolute_percentage_mode(self, col1: str, col2: str, threshold: float, highlight_mode: str = "entire_row", dividend_col: str = None) -> io.BytesIO:
        """
        Process absolute percentage difference mode - highlight above/below threshold with different colors

        Args:
            col1: First column name
            col2: Second column name  
            threshold: Threshold percentage
            highlight_mode: "entire_row" or "percentage_column"
            dividend_col: Column to use as dividend (denominator). If None, defaults to col2

        Returns:
            BytesIO object containing the Excel file
        """
        if self.df is None:
            raise ValueError("No data loaded")

        # Default dividend_col to col2 if not specified (maintains backward compatibility)
        if dividend_col is None:
            dividend_col = col2
        
        # Validate dividend_col
        if dividend_col not in [col1, col2]:
            raise ValueError(f"dividend_col must be either '{col1}' or '{col2}', got '{dividend_col}'")

        # Add placeholder column for percentage difference - will be replaced with formulas
        self.df['Percentage_Difference'] = 0

        return self._create_excel_output(mode=3, col1=col1, col2=col2, threshold=threshold, highlight_mode=highlight_mode, dividend_col=dividend_col)

    def _create_excel_output(self, mode: int, **kwargs) -> io.BytesIO:
        """
        Create Excel output with conditional formatting

        Args:
            mode: 1 for comparison, 2 for range
            **kwargs: Additional arguments based on mode

        Returns:
            BytesIO object containing the Excel file
        """
        output = io.BytesIO()

        # Write initial DataFrame to Excel
        self.df.to_excel(output, index=False, sheet_name=self.sheet_name)
        output.seek(0)

        wb_out = openpyxl.load_workbook(output)
        ws = wb_out[self.sheet_name]

        # Create Mask sheet
        mask_ws = wb_out.create_sheet("Mask")
        mask_ws['A1'] = 'HighlightMask'
        max_row = ws.max_row

        # Map headers to column letters
        headers = {cell.value: cell.column for cell in ws[1]}

        # Add dynamic formulas for percentage difference column if in comparison or absolute percentage mode
        if mode in [1, 3] and 'Percentage_Difference' in headers:
            col1, col2 = kwargs['col1'], kwargs['col2']
            dividend_col = kwargs.get('dividend_col', col2)  # Default to col2 for backward compatibility
            col1_letter = get_column_letter(headers[col1])
            col2_letter = get_column_letter(headers[col2])
            perc_diff_letter = get_column_letter(headers['Percentage_Difference'])
            
            # Determine dividend letter based on selected dividend column
            dividend_letter = col1_letter if dividend_col == col1 else col2_letter
            
            # Add dynamic Excel formulas for percentage difference
            for r in range(2, max_row + 1):
                formula = f"=ROUND((({col1_letter}{r}-{col2_letter}{r})/{dividend_letter}{r})*100,2)"
                ws[f"{perc_diff_letter}{r}"] = formula

        # Build mask formulas based on mode
        if mode == 1:
            # Comparison mode
            col1, col2, threshold = kwargs['col1'], kwargs['col2'], kwargs['threshold']
            col1_letter = get_column_letter(headers[col1])
            col2_letter = get_column_letter(headers[col2])

            for r in range(2, max_row + 1):
                formula = (
                    f"=ABS(({self.sheet_name}!{col1_letter}{r}-{self.sheet_name}!{col2_letter}{r})/{self.sheet_name}!{col2_letter}{r})*100<={threshold}"
                )
                mask_ws[f"A{r}"] = formula
        elif mode == 3:
            # Absolute percentage difference mode - create two mask columns
            col1, col2, threshold = kwargs['col1'], kwargs['col2'], kwargs['threshold']
            col1_letter = get_column_letter(headers[col1])
            col2_letter = get_column_letter(headers[col2])

            # Create second mask column header for "below threshold"
            mask_ws['B1'] = 'HighlightMask_Below'

            for r in range(2, max_row + 1):
                # Column A: Above threshold (faint red)
                formula_above = (
                    f"=ABS(({self.sheet_name}!{col1_letter}{r}-{self.sheet_name}!{col2_letter}{r})/{self.sheet_name}!{col2_letter}{r})*100>{threshold}"
                )
                mask_ws[f"A{r}"] = formula_above
                
                # Column B: Below or equal to threshold (faint blue)
                formula_below = (
                    f"=ABS(({self.sheet_name}!{col1_letter}{r}-{self.sheet_name}!{col2_letter}{r})/{self.sheet_name}!{col2_letter}{r})*100<={threshold}"
                )
                mask_ws[f"B{r}"] = formula_below
        else:
            # Range mode
            anchor_col, low_col, high_col = kwargs['anchor_col'], kwargs['low_col'], kwargs['high_col']
            anchor_letter = get_column_letter(headers[anchor_col])
            low_letter = get_column_letter(headers[low_col])
            high_letter = get_column_letter(headers[high_col])

            for r in range(2, max_row + 1):
                formula = f"=AND({self.sheet_name}!{anchor_letter}{r}>={self.sheet_name}!{low_letter}{r},{self.sheet_name}!{anchor_letter}{r}<={self.sheet_name}!{high_letter}{r})"
                mask_ws[f"A{r}"] = formula

        # Apply conditional formatting using mask sheet
        highlight_mode = kwargs.get('highlight_mode', 'entire_row')
        
        if mode == 3:
            # Absolute percentage difference mode - apply two different colored highlights
            red_fill = PatternFill(start_color='FFDDDD', end_color='FFDDDD', fill_type='solid')  # Faint red
            blue_fill = PatternFill(start_color='DDDDFF', end_color='DDDDFF', fill_type='solid')  # Faint blue
            
            if highlight_mode == 'entire_row':
                # Highlight entire rows
                first_col = get_column_letter(1)
                last_col = get_column_letter(ws.max_column)
                cell_range = f"{first_col}2:{last_col}{max_row}"
            else:
                # Highlight only percentage difference column
                if 'Percentage_Difference' in headers:
                    perc_diff_letter = get_column_letter(headers['Percentage_Difference'])
                    cell_range = f"{perc_diff_letter}2:{perc_diff_letter}{max_row}"
                else:
                    # Fallback to entire row if percentage column not found
                    first_col = get_column_letter(1)
                    last_col = get_column_letter(ws.max_column)
                    cell_range = f"{first_col}2:{last_col}{max_row}"
            
            # Apply red highlighting for above threshold
            cf_red = FormulaRule(formula=["=Mask!$A2=TRUE"], fill=red_fill)
            ws.conditional_formatting.add(cell_range, cf_red)
            
            # Apply blue highlighting for below threshold  
            cf_blue = FormulaRule(formula=["=Mask!$B2=TRUE"], fill=blue_fill)
            ws.conditional_formatting.add(cell_range, cf_blue)
        else:
            # Original single-color highlighting for modes 1 and 2
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            
            if highlight_mode == 'entire_row':
                # Highlight entire rows
                first_col = get_column_letter(1)
                last_col = get_column_letter(ws.max_column)
                cell_range = f"{first_col}2:{last_col}{max_row}"
            else:
                # Highlight only specific column based on mode
                if mode == 1:  # Comparison mode - highlight percentage difference column
                    if 'Percentage_Difference' in headers:
                        perc_diff_letter = get_column_letter(headers['Percentage_Difference'])
                        cell_range = f"{perc_diff_letter}2:{perc_diff_letter}{max_row}"
                    else:
                        # Fallback to entire row if percentage column not found
                        first_col = get_column_letter(1)
                        last_col = get_column_letter(ws.max_column)
                        cell_range = f"{first_col}2:{last_col}{max_row}"
                else:  # Range mode - highlight anchor column
                    anchor_col = kwargs.get('anchor_col')
                    if anchor_col and anchor_col in headers:
                        anchor_letter = get_column_letter(headers[anchor_col])
                        cell_range = f"{anchor_letter}2:{anchor_letter}{max_row}"
                    else:
                        # Fallback to entire row if anchor column not found
                        first_col = get_column_letter(1)
                        last_col = get_column_letter(ws.max_column)
                        cell_range = f"{first_col}2:{last_col}{max_row}"
            
            cf = FormulaRule(formula=["=Mask!$A2=TRUE"], fill=green_fill)
            ws.conditional_formatting.add(cell_range, cf)

        # Save to BytesIO
        output = io.BytesIO()
        wb_out.save(output)
        output.seek(0)

        return output
