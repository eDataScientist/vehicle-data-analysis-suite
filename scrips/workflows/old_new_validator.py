import pandas as pd
import io
import asyncio
from typing import Dict, List, Tuple, Optional, Union
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill


class OldNewDataValidator:
    """Multi-dimensional comparison between old and new vehicle data files"""

    def __init__(self):
        self.old_data = None
        self.new_data = None
        self.old_sheets = None
        self.new_sheets = None
        self.merged_data = None
        self.comparison_dimensions = [
            'ModelYear', 'Make', 'Model', 'Trim', 'BodyType', 
            'EngineSize', 'Transmission', 'Region', 'Doors', 'Seats', 'Cylinders'
        ]

    def load_old_data(self, file_data: Union[io.BytesIO, str], file_type: str = 'xlsx') -> List[str]:
        """Load old data file and return available sheets"""
        if file_type in ['xlsx', 'xls']:
            xls = pd.ExcelFile(file_data)
            self.old_sheets = xls.sheet_names
            return self.old_sheets
        else:
            self.old_data = pd.read_csv(file_data, encoding='utf-8')
            self.old_sheets = ['Data']
            return ['Data']

    def select_old_sheet(self, sheet_name: str, file_data: Union[io.BytesIO, str], file_type: str = 'xlsx') -> pd.DataFrame:
        """Select and load a specific sheet from old data"""
        if file_type in ['xlsx', 'xls']:
            xls = pd.ExcelFile(file_data)
            self.old_data = xls.parse(sheet_name)
        return self.old_data

    def load_new_data(self, file_data: Union[io.BytesIO, str], file_type: str = 'xlsx') -> List[str]:
        """Load new data file and return available sheets"""
        if file_type in ['xlsx', 'xls']:
            xls = pd.ExcelFile(file_data)
            self.new_sheets = xls.sheet_names
            return self.new_sheets
        else:
            self.new_data = pd.read_csv(file_data, encoding='utf-8')
            self.new_sheets = ['Data']
            return ['Data']

    def select_new_sheet(self, sheet_name: str, file_data: Union[io.BytesIO, str], file_type: str = 'xlsx') -> pd.DataFrame:
        """Select and load a specific sheet from new data"""
        if file_type in ['xlsx', 'xls']:
            xls = pd.ExcelFile(file_data)
            self.new_data = xls.parse(sheet_name)
        return self.new_data

    def get_old_columns(self) -> List[str]:
        """Get list of available columns from old data"""
        if self.old_data is not None:
            return list(self.old_data.columns)
        return []

    def get_new_columns(self) -> List[str]:
        """Get list of available columns from new data"""
        if self.new_data is not None:
            return list(self.new_data.columns)
        return []

    def prepare_old_data(self, column_mappings: Dict[str, str]) -> pd.DataFrame:
        """Prepare old data with column mappings"""
        if self.old_data is None:
            raise ValueError("Old data not loaded")
        
        # Check if columns exist in the data
        missing_cols = []
        for dimension, col_name in column_mappings.items():
            if col_name not in self.old_data.columns:
                missing_cols.append(f"{dimension}: '{col_name}'")
        
        if missing_cols:
            available_cols = list(self.old_data.columns)
            raise ValueError(f"Missing columns in old data: {missing_cols}. Available columns: {available_cols}")

        # Apply mappings and rename columns with _old suffix
        mapped_data = {'VIN': self.old_data['VIN']}  # Always include VIN for merging
        
        for dimension in self.comparison_dimensions:
            if dimension in column_mappings:
                mapped_data[f'{dimension}_old'] = self.old_data[column_mappings[dimension]]

        self.old_data = pd.DataFrame(mapped_data)
        return self.old_data

    def prepare_new_data(self, column_mappings: Dict[str, str]) -> pd.DataFrame:
        """Prepare new data with column mappings"""
        if self.new_data is None:
            raise ValueError("New data not loaded")
        
        # Check if columns exist in the data
        missing_cols = []
        for dimension, col_name in column_mappings.items():
            if col_name not in self.new_data.columns:
                missing_cols.append(f"{dimension}: '{col_name}'")
        
        if missing_cols:
            available_cols = list(self.new_data.columns)
            raise ValueError(f"Missing columns in new data: {missing_cols}. Available columns: {available_cols}")

        # Apply mappings and rename columns (no suffix for new data)
        mapped_data = {'VIN': self.new_data['VIN']}  # Always include VIN for merging
        
        for dimension in self.comparison_dimensions:
            if dimension in column_mappings:
                mapped_data[dimension] = self.new_data[column_mappings[dimension]]

        self.new_data = pd.DataFrame(mapped_data)
        return self.new_data

    def perform_multi_dimensional_comparison(self, old_column_mappings: Dict[str, str], new_column_mappings: Dict[str, str]) -> Dict[str, Union[int, float, pd.DataFrame]]:
        """Perform multi-dimensional comparison analysis between old and new data"""
        if self.old_data is None or self.new_data is None:
            raise ValueError("Both old and new data must be loaded")
        
        # First prepare the data with the provided column mappings
        self.prepare_old_data(old_column_mappings)
        self.prepare_new_data(new_column_mappings)
        
        # Create working copies
        old_working = self.old_data.copy()
        new_working = self.new_data.copy()

        # Merge datasets on VIN
        self.merged_data = new_working.merge(
            old_working,
            on='VIN',
            how='inner'  # Only compare records that exist in both datasets
        )

        # Calculate match statistics for each dimension
        match_results = {}
        total_records = len(self.merged_data)
        
        for dimension in self.comparison_dimensions:
            new_col = dimension
            old_col = f'{dimension}_old'
            
            # Skip dimensions not present in both datasets
            if new_col not in self.merged_data.columns or old_col not in self.merged_data.columns:
                continue
                
            # Handle different data types appropriately
            if dimension in ['ModelYear', 'Doors', 'Seats', 'Cylinders']:
                # Numeric comparison
                matches = (
                    self.merged_data[new_col].astype(str) == 
                    self.merged_data[old_col].astype(str)
                )
            else:
                # String comparison (case-insensitive)
                matches = (
                    self.merged_data[new_col].astype(str).str.lower() == 
                    self.merged_data[old_col].astype(str).str.lower()
                )
            
            match_count = matches.sum()
            match_percentage = match_count / total_records * 100 if total_records > 0 else 0
            
            match_results[dimension] = {
                'matches': match_count,
                'total': total_records,
                'percentage': match_percentage,
                'mismatches': total_records - match_count
            }

        # Calculate overall statistics
        overall_mismatches = self._get_overall_mismatches()
        
        return {
            'total_records': total_records,
            'dimension_results': match_results,
            'overall_mismatches': overall_mismatches,
            'overall_mismatches_count': len(overall_mismatches)
        }

    def _get_overall_mismatches(self) -> pd.DataFrame:
        """Get records that have mismatches in any dimension"""
        if self.merged_data is None:
            return pd.DataFrame()
        
        mismatch_mask = pd.Series([False] * len(self.merged_data))
        
        for dimension in self.comparison_dimensions:
            new_col = dimension
            old_col = f'{dimension}_old'
            
            if new_col not in self.merged_data.columns or old_col not in self.merged_data.columns:
                continue
                
            if dimension in ['ModelYear', 'Doors', 'Seats', 'Cylinders']:
                # Numeric comparison
                dimension_mismatch = (
                    self.merged_data[new_col].astype(str) != 
                    self.merged_data[old_col].astype(str)
                )
            else:
                # String comparison (case-insensitive)
                dimension_mismatch = (
                    self.merged_data[new_col].astype(str).str.lower() != 
                    self.merged_data[old_col].astype(str).str.lower()
                )
            
            mismatch_mask = mismatch_mask | dimension_mismatch
        
        return self.merged_data[mismatch_mask]

    def save_results(self) -> io.BytesIO:
        """Save analysis results to Excel file with dynamic conditional formatting"""
        if self.merged_data is None:
            raise ValueError("No comparison data available. Run comparison first.")

        output = io.BytesIO()

        # First save the main data
        self.merged_data.to_excel(output, sheet_name='Analysis', index=False)
        output.seek(0)

        # Load workbook for advanced formatting
        wb = openpyxl.load_workbook(output)
        ws = wb['Analysis']

        # Create dimension statistics for summary
        dimension_stats = []
        total_records = len(self.merged_data)
        
        for dimension in self.comparison_dimensions:
            new_col = dimension
            old_col = f'{dimension}_old'
            
            if new_col not in self.merged_data.columns or old_col not in self.merged_data.columns:
                continue
                
            if dimension in ['ModelYear', 'Doors', 'Seats', 'Cylinders']:
                matches = (
                    self.merged_data[new_col].astype(str) == 
                    self.merged_data[old_col].astype(str)
                )
            else:
                matches = (
                    self.merged_data[new_col].astype(str).str.lower() == 
                    self.merged_data[old_col].astype(str).str.lower()
                )
            
            match_count = matches.sum()
            match_percentage = match_count / total_records * 100 if total_records > 0 else 0
            
            dimension_stats.append({
                'Dimension': dimension,
                'Total_Records': total_records,
                'Matches': match_count,
                'Mismatches': total_records - match_count,
                'Match_Percentage': match_percentage
            })
        
        # Create Summary sheet
        summary_df = pd.DataFrame(dimension_stats)
        summary_ws = wb.create_sheet('Summary')
        
        # Write summary data
        for r_idx, row in enumerate(summary_df.values, 2):
            for c_idx, value in enumerate(row, 1):
                summary_ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Add headers to summary
        for c_idx, header in enumerate(summary_df.columns, 1):
            summary_ws.cell(row=1, column=c_idx, value=header)

        # Create Mismatches sheet
        overall_mismatches = self._get_overall_mismatches()
        if len(overall_mismatches) > 0:
            mismatch_ws = wb.create_sheet('Mismatches')
            for r_idx, row in enumerate(overall_mismatches.values, 2):
                for c_idx, value in enumerate(row, 1):
                    mismatch_ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Add headers to mismatches
            for c_idx, header in enumerate(overall_mismatches.columns, 1):
                mismatch_ws.cell(row=1, column=c_idx, value=header)

        # Apply dynamic conditional formatting to Analysis sheet
        self._apply_multi_dimensional_formatting(wb, ws)

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _apply_multi_dimensional_formatting(self, workbook, worksheet) -> None:
        """Apply dynamic conditional formatting for all 11 dimensions"""
        # Create Mask sheet for dynamic formulas
        mask_ws = workbook.create_sheet("Mask")
        
        # Add headers for mask columns A-K (11 dimensions)
        mask_headers = ['MakeMatch', 'ModelMatch', 'YearMatch', 'TrimMatch', 'BodyTypeMatch',
                       'EngineSizeMatch', 'TransmissionMatch', 'RegionMatch', 'DoorsMatch', 
                       'SeatsMatch', 'CylindersMatch']
        
        for i, header in enumerate(mask_headers, 1):
            mask_ws.cell(row=1, column=i, value=header)
        
        max_row = worksheet.max_row
        
        # Map headers to column letters
        headers = {cell.value: cell.column for cell in worksheet[1]}
        
        # Build mask formulas for each dimension
        mask_column_map = {}
        mask_col_idx = 1
        
        for dimension in self.comparison_dimensions:
            new_col_name = dimension
            old_col_name = f'{dimension}_old'
            
            if new_col_name in headers and old_col_name in headers:
                new_col_letter = get_column_letter(headers[new_col_name])
                old_col_letter = get_column_letter(headers[old_col_name])
                mask_col_letter = get_column_letter(mask_col_idx)
                
                # Build comparison formulas for each row
                for r in range(2, max_row + 1):
                    if dimension in ['ModelYear', 'Doors', 'Seats', 'Cylinders']:
                        # Numeric comparison
                        mask_ws[f"{mask_col_letter}{r}"] = f"=TEXT(Analysis!{new_col_letter}{r},\"0\")=TEXT(Analysis!{old_col_letter}{r},\"0\")"
                    else:
                        # String comparison (case-insensitive)
                        mask_ws[f"{mask_col_letter}{r}"] = f"=UPPER(Analysis!{new_col_letter}{r})=UPPER(Analysis!{old_col_letter}{r})"
                
                mask_column_map[old_col_name] = mask_col_letter
                mask_col_idx += 1
        
        # Define colors
        mint_green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        coral_pink = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
        
        # Apply conditional formatting to old columns (with _old suffix)
        for dimension in self.comparison_dimensions:
            old_col_name = f'{dimension}_old'
            
            if old_col_name in headers and old_col_name in mask_column_map:
                target_letter = get_column_letter(headers[old_col_name])
                mask_letter = mask_column_map[old_col_name]
                cell_range = f"{target_letter}2:{target_letter}{max_row}"
                
                # Green for matches
                green_rule = FormulaRule(formula=[f"=Mask!${mask_letter}2=TRUE"], fill=mint_green)
                worksheet.conditional_formatting.add(cell_range, green_rule)
                
                # Pink for non-matches
                pink_rule = FormulaRule(formula=[f"=Mask!${mask_letter}2=FALSE"], fill=coral_pink)
                worksheet.conditional_formatting.add(cell_range, pink_rule)

    def get_comparison_summary(self) -> Dict[str, Union[int, float]]:
        """Get summary statistics from multi-dimensional comparison"""
        if self.merged_data is None:
            return {}

        summary = {'total_records': len(self.merged_data)}
        
        for dimension in self.comparison_dimensions:
            new_col = dimension
            old_col = f'{dimension}_old'
            
            if new_col not in self.merged_data.columns or old_col not in self.merged_data.columns:
                continue
                
            if dimension in ['ModelYear', 'Doors', 'Seats', 'Cylinders']:
                matches = (
                    self.merged_data[new_col].astype(str) == 
                    self.merged_data[old_col].astype(str)
                )
            else:
                matches = (
                    self.merged_data[new_col].astype(str).str.lower() == 
                    self.merged_data[old_col].astype(str).str.lower()
                )
            
            match_count = matches.sum()
            total = len(self.merged_data)
            
            summary[f'{dimension}_matches'] = match_count
            summary[f'{dimension}_match_percentage'] = match_count / total * 100 if total > 0 else 0

        return summary

    def get_sample_mismatches(self, n: int = 5) -> pd.DataFrame:
        """Get sample mismatches for display"""
        overall_mismatches = self._get_overall_mismatches()
        
        if len(overall_mismatches) == 0:
            return pd.DataFrame()

        # Select key columns for display
        display_cols = ['VIN']
        for dimension in self.comparison_dimensions[:6]:  # Show first 6 dimensions
            new_col = dimension
            old_col = f'{dimension}_old'
            if new_col in overall_mismatches.columns and old_col in overall_mismatches.columns:
                display_cols.extend([new_col, old_col])

        available_cols = [col for col in display_cols if col in overall_mismatches.columns]
        return overall_mismatches[available_cols].head(n)