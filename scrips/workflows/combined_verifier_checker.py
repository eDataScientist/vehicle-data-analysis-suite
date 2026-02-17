import pandas as pd
import io
import asyncio
from typing import Dict, List, Tuple, Optional, Union
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule

# Import existing workflows
from .vehicle_verifier import VehicleDataVerifier
from .discrepancy_checker import DiscrepancyChecker


class CombinedVerifierChecker:
    """Combined Vehicle Data Verifier and Discrepancy Checker workflow - Excel-based approach"""

    def __init__(self):
        # Initialize vehicle verifier component
        self.verifier = VehicleDataVerifier()
        self.analysis_results = None

        # Store configuration for discrepancy analysis
        self.discrepancy_config = None

        # Store FULL original data before verifier processes it
        self.full_reference_data = None
        self.full_logs_data = None
        self.extra_match_cols: List[Tuple[str, str]] = []

    # Delegate file operations to verifier
    def load_reference_data(
        self, file_data: Union[io.BytesIO, str], file_type: str = "xlsx"
    ) -> List[str]:
        return self.verifier.load_reference_data(file_data, file_type)

    def select_reference_sheet(
        self,
        sheet_name: str,
        file_data: Union[io.BytesIO, str],
        file_type: str = "xlsx",
    ) -> pd.DataFrame:
        df = self.verifier.select_reference_sheet(sheet_name, file_data, file_type)
        # Store FULL original data with ALL columns
        self.full_reference_data = df.copy()
        return df

    def load_logs_data(
        self, file_data: Union[io.BytesIO, str], file_type: str = "xlsx"
    ) -> List[str]:
        return self.verifier.load_logs_data(file_data, file_type)

    def select_logs_sheet(
        self,
        sheet_name: str,
        file_data: Union[io.BytesIO, str],
        file_type: str = "xlsx",
    ) -> pd.DataFrame:
        df = self.verifier.select_logs_sheet(sheet_name, file_data, file_type)
        # Store FULL original data with ALL columns
        self.full_logs_data = df.copy()
        return df

    def get_reference_columns(self) -> List[str]:
        return self.verifier.get_reference_columns()

    def get_logs_columns(self) -> List[str]:
        return self.verifier.get_logs_columns()

    async def perform_translation(
        self, api_key: str, progress_callback=None
    ) -> Dict[str, str]:
        return await self.verifier.perform_translation(api_key, progress_callback)

    def perform_combined_analysis(
        self,
        # Vehicle verification parameters
        chassis_col: str,
        make_ext_col: str,
        model_ext_col: str,
        year_ext_col: str,
        vin_col: str,
        make_col: str,
        model_col: str,
        year_col: str,
        spec_status_col: str,
        # Discrepancy analysis parameters
        analysis_mode: str,
        val_col1: str = None,
        val_col2: str = None,
        threshold: float = 15.0,
        dividend_col: str = None,
        highlight_mode: str = "entire_row",
        anchor_col: str = None,
        low_col: str = None,
        high_col: str = None,
        # Extra matching fields
        extra_match_cols: List[Tuple[str, str]] = None,
    ) -> Dict:
        """Perform combined analysis by working directly with Excel workbook"""
        self.extra_match_cols = extra_match_cols if extra_match_cols else []

        # Perform vehicle verification and get statistics
        verification_results = self.verifier.perform_verification(
            chassis_col,
            make_ext_col,
            model_ext_col,
            year_ext_col,
            vin_col,
            make_col,
            model_col,
            year_col,
            spec_status_col,
            extra_match_cols=extra_match_cols,
        )

        # Store discrepancy configuration for later use
        self.discrepancy_config = {
            "analysis_mode": analysis_mode,
            "val_col1": val_col1,
            "val_col2": val_col2,
            "threshold": threshold,
            "dividend_col": dividend_col,
            "highlight_mode": highlight_mode,
            "anchor_col": anchor_col,
            "low_col": low_col,
            "high_col": high_col,
        }

        # Store results
        self.analysis_results = {
            "verification_results": verification_results,
            "analysis_mode": analysis_mode,
        }

        return self.analysis_results

    def _resolve_header_aliases(self, col_name: str) -> List[str]:
        """Resolve possible header aliases for merged output columns."""
        aliases = [col_name, f"{col_name}_primary", f"{col_name}_join"]

        # If this is an extra matched pair, prefer side-specific aliases.
        for ref_col, logs_col in self.extra_match_cols:
            if col_name == logs_col:
                aliases = [logs_col, f"{logs_col}_primary", f"{logs_col}_join"] + [
                    a for a in aliases if a not in {logs_col, f"{logs_col}_primary", f"{logs_col}_join"}
                ]
                break
            if col_name == ref_col:
                aliases = [ref_col, f"{ref_col}_join", f"{ref_col}_primary"] + [
                    a for a in aliases if a not in {ref_col, f"{ref_col}_join", f"{ref_col}_primary"}
                ]
                break

        # Preserve order and uniqueness
        seen = set()
        ordered_aliases = []
        for alias in aliases:
            if alias not in seen:
                seen.add(alias)
                ordered_aliases.append(alias)
        return ordered_aliases

    def _find_worksheet_column_idx(self, worksheet, col_name: str, header_row: int = 1):
        """Find worksheet column index using suffix-aware header aliases."""
        header_to_idx = {
            worksheet.cell(row=header_row, column=col_idx).value: col_idx
            for col_idx in range(1, worksheet.max_column + 1)
        }
        for alias in self._resolve_header_aliases(col_name):
            if alias in header_to_idx:
                return header_to_idx[alias]
        return None

    def save_combined_results(self, include_mask_in_main: bool = True) -> io.BytesIO:
        """Save combined results by working directly with Excel workbook"""
        if self.analysis_results is None:
            raise ValueError(
                "No analysis results available. Run combined analysis first."
            )

        try:
            # Get the Excel file from verifier (with formulas and formatting intact)
            verifier_excel = self.verifier.save_results(
                include_mask_in_main=include_mask_in_main
            )

            # If no discrepancy config, just return the verifier results
            if self.discrepancy_config is None:
                return verifier_excel

            # Load the Excel workbook with openpyxl
            verifier_excel.seek(0)
            workbook = load_workbook(verifier_excel)

            # Work with the main analysis sheet
            if "Verification Analysis" in workbook.sheetnames:
                worksheet = workbook["Verification Analysis"]
            else:
                worksheet = workbook.active

            # Get the merged data
            merged_data = self.verifier.merged_data

            # Apply discrepancy analysis based on mode
            config = self.discrepancy_config
            analysis_mode = config["analysis_mode"]

            if analysis_mode in [
                "Compare two columns within % threshold",
                "Absolute % difference with color coding",
            ]:
                self._add_percentage_discrepancy_to_excel(
                    worksheet, merged_data, config, include_mask_in_main
                )
            elif analysis_mode == "Check value between low & high columns":
                self._add_range_check_to_excel(
                    worksheet, merged_data, config, include_mask_in_main
                )

            # Save the modified workbook
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            return output

        except Exception as e:
            print(f"Error in save_combined_results: {str(e)}")
            import traceback

            print(traceback.format_exc())
            # Fallback to verifier results
            return self.verifier.save_results(include_mask_in_main=include_mask_in_main)

    def _add_percentage_discrepancy_to_excel(
        self,
        worksheet,
        merged_data: pd.DataFrame,
        config: Dict,
        include_mask_in_main: bool,
    ):
        """Add percentage discrepancy columns and formatting to Excel worksheet"""
        val_col1 = config["val_col1"]
        val_col2 = config["val_col2"]
        threshold = config["threshold"]
        dividend_col = config["dividend_col"]
        analysis_mode = config["analysis_mode"]

        # Find the header row and last column
        header_row = 1
        last_col_idx = worksheet.max_column

        # Find column indices for val_col1 and val_col2 in the Excel sheet
        val_col1_idx = self._find_worksheet_column_idx(worksheet, val_col1, header_row)
        val_col2_idx = self._find_worksheet_column_idx(worksheet, val_col2, header_row)

        # Append missing columns from full original data
        if val_col1_idx is None:
            col_data = self._find_column_data(val_col1, merged_data)
            if col_data is not None:
                last_col_idx += 1
                val_col1_idx = last_col_idx
                worksheet.cell(row=header_row, column=val_col1_idx, value=val_col1)
                for row_idx, value in enumerate(col_data, start=2):
                    worksheet.cell(row=row_idx, column=val_col1_idx, value=value)
            else:
                raise ValueError(f"Column '{val_col1}' not found in any data source")

        if val_col2_idx is None:
            col_data = self._find_column_data(val_col2, merged_data)
            if col_data is not None:
                last_col_idx += 1
                val_col2_idx = last_col_idx
                worksheet.cell(row=header_row, column=val_col2_idx, value=val_col2)
                for row_idx, value in enumerate(col_data, start=2):
                    worksheet.cell(row=row_idx, column=val_col2_idx, value=value)
            else:
                raise ValueError(f"Column '{val_col2}' not found in any data source")

        # Convert column indices to Excel column letters
        from openpyxl.utils import get_column_letter

        val_col1_letter = get_column_letter(val_col1_idx)
        val_col2_letter = get_column_letter(val_col2_idx)

        # Add percentage difference column
        last_col_idx += 1
        percent_diff_col_idx = last_col_idx
        percent_diff_col_letter = get_column_letter(percent_diff_col_idx)

        # Set header
        worksheet.cell(
            row=header_row,
            column=percent_diff_col_idx,
            value="Percentage Difference (%)",
        )

        # Add formulas for each row
        total_rows = len(merged_data) + 1  # +1 for header
        for row_idx in range(2, total_rows + 1):
            # Create formula based on dividend column
            if dividend_col == val_col1:
                dividend_col_letter = val_col1_letter
            else:
                dividend_col_letter = val_col2_letter

            formula = f'=IF(AND(ISNUMBER({val_col1_letter}{row_idx}),ISNUMBER({val_col2_letter}{row_idx})),IF({dividend_col_letter}{row_idx}=0,"Division by Zero",(({val_col1_letter}{row_idx}-{val_col2_letter}{row_idx})/{dividend_col_letter}{row_idx})*100),"Invalid Data")'
            # formula = f'=IF(ISNUMBER({val_col2_letter}{row_idx}),IF({val_col2_letter}{row_idx}=0,"Division by Zero",(({val_col1_letter}{row_idx}-{val_col2_letter}{row_idx})/{val_col2_letter}{row_idx})*100),"Invalid Data")'

            worksheet.cell(row=row_idx, column=percent_diff_col_idx, value=formula)

        # Apply conditional formatting
        if analysis_mode == "Compare two columns within % threshold":
            # Red fill for values exceeding threshold (absolute value)
            red_fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )
            red_font = Font(color="9C0006")

            # Format range for percentage difference column
            range_address = (
                f"{percent_diff_col_letter}2:{percent_diff_col_letter}{total_rows}"
            )

            # Add conditional formatting rule for values > threshold
            worksheet.conditional_formatting.add(
                range_address,
                CellIsRule(
                    operator="greaterThan",
                    formula=[str(threshold)],
                    fill=red_fill,
                    font=red_font,
                ),
            )

            # Add conditional formatting rule for values < -threshold
            worksheet.conditional_formatting.add(
                range_address,
                CellIsRule(
                    operator="lessThan",
                    formula=[str(-threshold)],
                    fill=red_fill,
                    font=red_font,
                ),
            )

        elif analysis_mode == "Absolute % difference with color coding":
            # Faint red for values above threshold
            red_fill = PatternFill(
                start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
            )
            # Faint blue for values below negative threshold
            blue_fill = PatternFill(
                start_color="E6F2FF", end_color="E6F2FF", fill_type="solid"
            )

            range_address = (
                f"{percent_diff_col_letter}2:{percent_diff_col_letter}{total_rows}"
            )

            # Red for positive values > threshold
            worksheet.conditional_formatting.add(
                range_address,
                CellIsRule(
                    operator="greaterThan", formula=[str(threshold)], fill=red_fill
                ),
            )

            # Blue for negative values < -threshold
            worksheet.conditional_formatting.add(
                range_address,
                CellIsRule(
                    operator="lessThan", formula=[str(-threshold)], fill=blue_fill
                ),
            )

    def _add_range_check_to_excel(
        self,
        worksheet,
        merged_data: pd.DataFrame,
        config: Dict,
        include_mask_in_main: bool,
    ):
        """Add range check columns and formatting to Excel worksheet"""
        anchor_col = config["anchor_col"]
        low_col = config["low_col"]
        high_col = config["high_col"]

        # Find the header row and last column
        header_row = 1
        last_col_idx = worksheet.max_column

        # Find column indices
        anchor_col_idx = self._find_worksheet_column_idx(worksheet, anchor_col, header_row)
        low_col_idx = self._find_worksheet_column_idx(worksheet, low_col, header_row)
        high_col_idx = self._find_worksheet_column_idx(worksheet, high_col, header_row)

        # Append missing columns from full original data
        if anchor_col_idx is None:
            col_data = self._find_column_data(anchor_col, merged_data)
            if col_data is not None:
                last_col_idx += 1
                anchor_col_idx = last_col_idx
                worksheet.cell(row=header_row, column=anchor_col_idx, value=anchor_col)
                for row_idx, value in enumerate(col_data, start=2):
                    worksheet.cell(row=row_idx, column=anchor_col_idx, value=value)
            else:
                raise ValueError(f"Column '{anchor_col}' not found in any data source")

        if low_col_idx is None:
            col_data = self._find_column_data(low_col, merged_data)
            if col_data is not None:
                last_col_idx += 1
                low_col_idx = last_col_idx
                worksheet.cell(row=header_row, column=low_col_idx, value=low_col)
                for row_idx, value in enumerate(col_data, start=2):
                    worksheet.cell(row=row_idx, column=low_col_idx, value=value)
            else:
                raise ValueError(f"Column '{low_col}' not found in any data source")

        if high_col_idx is None:
            col_data = self._find_column_data(high_col, merged_data)
            if col_data is not None:
                last_col_idx += 1
                high_col_idx = last_col_idx
                worksheet.cell(row=header_row, column=high_col_idx, value=high_col)
                for row_idx, value in enumerate(col_data, start=2):
                    worksheet.cell(row=row_idx, column=high_col_idx, value=value)
            else:
                raise ValueError(f"Column '{high_col}' not found in any data source")

        # Convert to column letters
        from openpyxl.utils import get_column_letter

        anchor_letter = get_column_letter(anchor_col_idx)
        low_letter = get_column_letter(low_col_idx)
        high_letter = get_column_letter(high_col_idx)

        # Add "Within Range" column
        last_col_idx += 1
        within_range_col_idx = last_col_idx
        within_range_letter = get_column_letter(within_range_col_idx)

        worksheet.cell(
            row=header_row, column=within_range_col_idx, value="Within Range"
        )

        # Add formulas
        total_rows = len(merged_data) + 1
        for row_idx in range(2, total_rows + 1):
            formula = f"=AND({anchor_letter}{row_idx}>={low_letter}{row_idx},{anchor_letter}{row_idx}<={high_letter}{row_idx})"
            worksheet.cell(row=row_idx, column=within_range_col_idx, value=formula)

        # Apply conditional formatting - red fill for FALSE (out of range)
        red_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        red_font = Font(color="9C0006")

        range_address = f"{within_range_letter}2:{within_range_letter}{total_rows}"

        worksheet.conditional_formatting.add(
            range_address,
            CellIsRule(
                operator="equal", formula=["FALSE"], fill=red_fill, font=red_font
            ),
        )

    def _find_column_data(self, col_name: str, merged_data: pd.DataFrame):
        """Helper to find column data from various sources"""
        for alias in self._resolve_header_aliases(col_name):
            if alias in merged_data.columns:
                return merged_data[alias]
        if (
            self.full_reference_data is not None
            and col_name in self.full_reference_data.columns
        ):
            return self.full_reference_data[col_name]
        if (
            self.full_logs_data is not None and col_name in self.full_logs_data.columns
        ):
            return self.full_logs_data[col_name]
        return None

    def get_merged_data(self) -> pd.DataFrame:
        """Get the merged data from verifier"""
        return self.verifier.merged_data

    def get_sample_mismatches(self, n: int = 5) -> pd.DataFrame:
        """Get sample mismatches using verifier's method"""
        return self.verifier.get_sample_mismatches(n)

    def get_verification_summary(self) -> Dict[str, Union[int, float]]:
        """Get verification summary using verifier's method"""
        return self.verifier.get_verification_summary()
