import pandas as pd
import io
import asyncio
from typing import Dict, List, Tuple, Optional, Union
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

# Import translation service
try:
    from ..services.translation_service import ArabicTranslationService

    TRANSLATION_AVAILABLE = True
except ImportError:
    TRANSLATION_AVAILABLE = False


class VehicleDataVerifier:
    """Vehicle data verification with optional translation capabilities"""

    def __init__(self):
        self.reference_data = None
        self.logs_data = None
        self.reference_sheets = None
        self.logs_sheets = None
        self.merged_data = None
        # Initialize translation service if available
        self.translation_service = (
            ArabicTranslationService() if TRANSLATION_AVAILABLE else None
        )

    def load_reference_data(
        self, file_data: Union[io.BytesIO, str], file_type: str = "xlsx"
    ) -> List[str]:
        """Load reference data file and return available sheets"""
        if file_type in ["xlsx", "xls"]:
            xls = pd.ExcelFile(file_data)
            self.reference_sheets = xls.sheet_names
            return self.reference_sheets
        else:
            self.reference_data = pd.read_csv(file_data, encoding="utf-8")
            self.reference_sheets = ["Data"]
            return ["Data"]

    def select_reference_sheet(
        self,
        sheet_name: str,
        file_data: Union[io.BytesIO, str],
        file_type: str = "xlsx",
    ) -> pd.DataFrame:
        """Select and load a specific sheet from reference data"""
        if file_type in ["xlsx", "xls"]:
            xls = pd.ExcelFile(file_data)
            self.reference_data = xls.parse(sheet_name)
        return self.reference_data

    def load_logs_data(
        self, file_data: Union[io.BytesIO, str], file_type: str = "xlsx"
    ) -> List[str]:
        """Load logs data file and return available sheets"""
        if file_type in ["xlsx", "xls"]:
            xls = pd.ExcelFile(file_data)
            self.logs_sheets = xls.sheet_names
            return self.logs_sheets
        else:
            self.logs_data = pd.read_csv(file_data, encoding="utf-8")
            self.logs_sheets = ["Data"]
            return ["Data"]

    def select_logs_sheet(
        self,
        sheet_name: str,
        file_data: Union[io.BytesIO, str],
        file_type: str = "xlsx",
    ) -> pd.DataFrame:
        """Select and load a specific sheet from logs data"""
        if file_type in ["xlsx", "xls"]:
            xls = pd.ExcelFile(file_data)
            self.logs_data = xls.parse(sheet_name)
        return self.logs_data

    def _resolve_merged_column(
        self, col_name: str, merged_cols: List[str], suffix: str
    ) -> str:
        """Resolve a column name in merged data, checking for suffixed variants.

        After a pandas merge with suffixes, columns that exist in both DataFrames
        get suffixed. This method checks for the suffixed version first, then
        the raw name, and raises a clear error if neither is found.
        """
        suffixed = col_name + suffix
        if suffixed in merged_cols:
            return suffixed
        if col_name in merged_cols:
            return col_name
        raise KeyError(
            f"Column '{col_name}' not found in merged data (also tried '{suffixed}'). "
            f"Available columns: {merged_cols}"
        )

    def get_reference_columns(self) -> List[str]:
        """Get list of available columns from reference data"""
        if self.reference_data is not None:
            return list(self.reference_data.columns)
        return []

    def get_logs_columns(self) -> List[str]:
        """Get list of available columns from logs data"""
        if self.logs_data is not None:
            return list(self.logs_data.columns)
        return []

    def prepare_reference_data(
        self,
        chassis_col: str,
        make_col: str,
        model_col: str,
        year_col: str,
        extra_cols: List[str] = None,
    ) -> pd.DataFrame:
        """Prepare reference data with column mappings

        Args:
            chassis_col: Chassis column name
            make_col: Make column name
            model_col: Model column name
            year_col: Year column name
            extra_cols: Additional columns to include in the output
        """
        if self.reference_data is None:
            raise ValueError("Reference data not loaded")

        # Check if required columns exist in the data
        missing_cols = []
        for col_name, col_val in [
            ("chassis_col", chassis_col),
            ("make_col", make_col),
            ("model_col", model_col),
            ("year_col", year_col),
        ]:
            if col_val not in self.reference_data.columns:
                missing_cols.append(f"{col_name}: '{col_val}'")

        if missing_cols:
            available_cols = list(self.reference_data.columns)
            raise ValueError(
                f"Missing columns in reference data: {missing_cols}. Available columns: {available_cols}"
            )

        # Apply mappings and rename columns
        mapped_data = {
            "chassis_no": self.reference_data[chassis_col],
            "Make_ext": self.reference_data[make_col],
            "Model_ext": self.reference_data[model_col],
            "ModelYear_ext": self.reference_data[year_col],
        }

        # Add extra columns with original names
        if extra_cols:
            missing_extra = []
            for col in extra_cols:
                if col in self.reference_data.columns:
                    mapped_data[col] = self.reference_data[col]
                else:
                    missing_extra.append(col)
            if missing_extra:
                available_cols = list(self.reference_data.columns)
                raise ValueError(
                    f"Extra columns not found in reference data: {missing_extra}. "
                    f"Available columns: {available_cols}"
                )

        self.reference_data = pd.DataFrame(mapped_data)
        return self.reference_data

    def prepare_logs_data(
        self,
        vin_col: str,
        make_col: str,
        model_col: str,
        year_col: str,
        spec_status_col: str,
        extra_cols: List[str] = None,
    ) -> pd.DataFrame:
        """Prepare logs data with column mappings

        Args:
            vin_col: VIN column name
            make_col: Make column name
            model_col: Model column name
            year_col: Year column name
            spec_status_col: Specification status column name
            extra_cols: Additional columns to include in the output
        """
        if self.logs_data is None:
            raise ValueError("Logs data not loaded")

        # Check if required columns exist in the data
        missing_cols = []

        for col_name, col_val in [
            ("vin_col", vin_col),
            ("make_col", make_col),
            ("model_col", model_col),
            ("year_col", year_col),
            ("spec_status_col", spec_status_col),
        ]:
            if col_val not in self.logs_data.columns:
                missing_cols.append(f"{col_name}: '{col_val}'")

        if missing_cols:
            available_cols = list(self.logs_data.columns)
            raise ValueError(
                f"Missing columns in logs data: {missing_cols}. Available columns: {available_cols}"
            )

        # Apply mappings
        mapped_data = {
            "VIN": self.logs_data[vin_col],
            "Make": self.logs_data[make_col],
            "Model": self.logs_data[model_col],
            "ModelYear": self.logs_data[year_col],
            "Specification Status": self.logs_data[spec_status_col],
        }

        # Add extra columns with original names
        if extra_cols:
            missing_extra = []
            for col in extra_cols:
                if col in self.logs_data.columns:
                    mapped_data[col] = self.logs_data[col]
                else:
                    missing_extra.append(col)
            if missing_extra:
                available_cols = list(self.logs_data.columns)
                raise ValueError(
                    f"Extra columns not found in logs data: {missing_extra}. "
                    f"Available columns: {available_cols}"
                )

        self.logs_data = pd.DataFrame(mapped_data)
        return self.logs_data

    async def perform_translation(
        self, api_key: str, progress_callback=None
    ) -> Dict[str, str]:
        """Perform Arabic to English translation on brand and model names"""
        if not TRANSLATION_AVAILABLE or not self.translation_service:
            return {"error": "Translation not available. Install openai library."}

        if self.reference_data is None:
            return {"error": "Reference data not loaded"}

        try:
            # Use the translation service to translate the DataFrame columns
            self.reference_data = (
                await self.translation_service.translate_dataframe_columns(
                    self.reference_data,
                    ["Make_ext", "Model_ext"],
                    api_key,
                    progress_callback,
                )
            )

            # Get statistics for response
            brand_strings = self.reference_data["Make_ext"].astype(str).tolist()
            model_strings = self.reference_data["Model_ext"].astype(str).tolist()

            ar_brand_count = len(
                [
                    s
                    for s in brand_strings
                    if self.translation_service.detect_arabic_text(s)
                ]
            )
            ar_model_count = len(
                [
                    s
                    for s in model_strings
                    if self.translation_service.detect_arabic_text(s)
                ]
            )

            if ar_brand_count == 0 and ar_model_count == 0:
                return {"status": "Translation completed - no Arabic text remaining"}

            return {
                "status": "Translation completed",
                "brand_translations": ar_brand_count,
                "model_translations": ar_model_count,
            }
        except Exception as e:
            return {"error": f"Translation failed: {str(e)}"}

    def perform_verification(
        self,
        chassis_col: str,
        make_ext_col: str,
        model_ext_col: str,
        year_ext_col: str,
        vin_col: str,
        make_col: str,
        model_col: str,
        year_col: str,
        spec_status_col: str,
        extra_match_cols: List[Tuple[str, str]] = None,  # type: ignore[assignment]
    ) -> Dict[str, Union[int, float, pd.DataFrame]]:
        """Perform verification analysis between reference and logs data

        Args:
            chassis_col: Chassis column in reference data
            make_ext_col: Make column in reference data
            model_ext_col: Model column in reference data
            year_ext_col: Year column in reference data
            vin_col: VIN column in logs data
            make_col: Make column in logs data
            model_col: Model column in logs data
            year_col: Year column in logs data
            spec_status_col: Specification status column in logs data
            extra_match_cols: List of tuples (ref_col, logs_col) for additional field matching

        Returns:
            Dictionary with verification statistics and mismatches
        """
        if self.reference_data is None or self.logs_data is None:
            raise ValueError("Both reference and logs data must be loaded")

        # Store extra match cols for later use
        self.extra_match_cols = extra_match_cols if extra_match_cols else []

        # Extract extra column names from extra_match_cols
        extra_ref_cols = (
            [col[0] for col in self.extra_match_cols] if self.extra_match_cols else None
        )
        extra_logs_cols = (
            [col[1] for col in self.extra_match_cols] if self.extra_match_cols else None
        )

        # First prepare the data with the provided column mappings
        self.prepare_reference_data(
            chassis_col,
            make_ext_col,
            model_ext_col,
            year_ext_col,
            extra_cols=extra_ref_cols,
        )

        self.prepare_logs_data(
            vin_col,
            make_col,
            model_col,
            year_col,
            spec_status_col,
            extra_cols=extra_logs_cols,
        )

        # Now use the standardized column names for the working copies
        ref_working = self.reference_data.copy()
        logs_working = self.logs_data.copy()

        # Merge datasets using standardized column names
        self.merged_data = logs_working.merge(
            ref_working,
            left_on="VIN",
            right_on="chassis_no",
            how="left",
            suffixes=("_primary", "_join"),
        )

        # Calculate match statistics for reporting (but don't add to merged_data)
        make_matches_calc = (
            self.merged_data["Make"].str.lower()
            == self.merged_data["Make_ext"].str.lower()
        )

        model_matches_calc = (
            self.merged_data["Model"].str.lower()
            == self.merged_data["Model_ext"].str.lower()
        )

        year_matches_calc = self.merged_data["ModelYear"].astype(
            str
        ) == self.merged_data["ModelYear_ext"].astype(str)

        # Calculate extra field matches
        extra_matches = {}
        if self.extra_match_cols:
            merged_cols = list(self.merged_data.columns)
            for ref_col, logs_col in self.extra_match_cols:
                # Find the actual column names in merged data
                # When both sides share the same column name, pandas adds suffixes:
                # logs (left) gets _primary, reference (right) gets _join
                # When only one side has the column, no suffix is added
                ref_col_actual = self._resolve_merged_column(
                    ref_col, merged_cols, "_join"
                )
                logs_col_actual = self._resolve_merged_column(
                    logs_col, merged_cols, "_primary"
                )

                ref_vals = self.merged_data[ref_col_actual].astype(str).str.lower()
                logs_vals = (
                    self.merged_data[logs_col_actual].astype(str).str.lower()
                )
                extra_matches[f"{logs_col}_match"] = ref_vals == logs_vals

        # Calculate match statistics
        total_records = len(self.merged_data)
        make_matches = make_matches_calc.sum()
        model_matches = model_matches_calc.sum()
        year_matches = year_matches_calc.sum()

        # Calculate extra field match counts
        extra_match_counts = {}
        for match_name, match_series in extra_matches.items():
            extra_match_counts[match_name] = match_series.sum()

        # Build mismatch condition - start with Make/Model/Year
        mismatch_condition = (
            ~make_matches_calc | ~model_matches_calc | ~year_matches_calc
        )
        # Add extra field mismatches
        for match_name, match_series in extra_matches.items():
            mismatch_condition = mismatch_condition | ~match_series

        # Get mismatches
        mismatches = self.merged_data[mismatch_condition]

        # Build return dictionary
        result = {
            "total_records": total_records,
            "make_matches": make_matches,
            "model_matches": model_matches,
            "year_matches": year_matches,
            "make_match_percentage": make_matches / total_records * 100
            if total_records > 0
            else 0,
            "model_match_percentage": model_matches / total_records * 100
            if total_records > 0
            else 0,
            "year_match_percentage": year_matches / total_records * 100
            if total_records > 0
            else 0,
            "mismatches": mismatches,
            "mismatches_count": len(mismatches),
        }

        # Add extra field match info
        result["extra_match_counts"] = extra_match_counts
        result["extra_match_percentages"] = {
            k: v / total_records * 100 if total_records > 0 else 0
            for k, v in extra_match_counts.items()
        }

        return result

    def save_results(self, include_mask_in_main: bool = True) -> io.BytesIO:
        """Save analysis results to Excel file with dynamic conditional formatting"""
        if self.merged_data is None:
            raise ValueError("No verification data available. Run verification first.")

        # Get extra_match_cols (stored during perform_verification)
        extra_match_cols = getattr(self, "extra_match_cols", [])

        # Calculate match statistics
        make_matches_calc = (
            self.merged_data["Make"].str.lower()
            == self.merged_data["Make_ext"].str.lower()
        )
        model_matches_calc = (
            self.merged_data["Model"].str.lower()
            == self.merged_data["Model_ext"].str.lower()
        )
        year_matches_calc = self.merged_data["ModelYear"].astype(
            str
        ) == self.merged_data["ModelYear_ext"].astype(str)

        # Calculate extra field matches
        extra_matches = {}
        extra_match_cols = getattr(self, "extra_match_cols", [])
        if extra_match_cols:
            merged_cols = list(self.merged_data.columns)
            for ref_col, logs_col in extra_match_cols:
                match_name = f"{logs_col}_match"
                ref_col_actual = self._resolve_merged_column(
                    ref_col, merged_cols, "_join"
                )
                logs_col_actual = self._resolve_merged_column(
                    logs_col, merged_cols, "_primary"
                )
                ref_vals = self.merged_data[ref_col_actual].astype(str).str.lower()
                logs_vals = self.merged_data[logs_col_actual].astype(str).str.lower()
                extra_matches[match_name] = ref_vals == logs_vals

        # Prepare data for saving
        data_to_save = self.merged_data.copy()

        # Add match columns to main sheet if requested
        if include_mask_in_main:
            data_to_save["Make Match"] = make_matches_calc
            data_to_save["Model Match"] = model_matches_calc
            data_to_save["Year Match"] = year_matches_calc
            # Add extra field match columns
            for match_name, match_series in extra_matches.items():
                data_to_save[match_name] = match_series

        output = io.BytesIO()

        # Save the main data
        data_to_save.to_excel(output, sheet_name="Analysis", index=False)
        output.seek(0)

        # Load workbook for advanced formatting
        wb = openpyxl.load_workbook(output)
        ws = wb["Analysis"]

        # Replace boolean values with Excel formulas if match columns are included
        if include_mask_in_main:
            # Map headers to column letters
            headers = {cell.value: cell.column for cell in ws[1]}

            # Get column letters for data columns
            make_col = get_column_letter(headers.get("Make", 1))
            make_ext_col = get_column_letter(headers.get("Make_ext", 1))
            model_col = get_column_letter(headers.get("Model", 1))
            model_ext_col = get_column_letter(headers.get("Model_ext", 1))
            year_col = get_column_letter(headers.get("ModelYear", 1))
            year_ext_col = get_column_letter(headers.get("ModelYear_ext", 1))

            max_row = ws.max_row

            # Replace boolean values with Excel formulas in match columns
            if "Make Match" in headers:
                make_match_col = get_column_letter(headers["Make Match"])
                for r in range(2, max_row + 1):
                    ws[f"{make_match_col}{r}"] = (
                        f"=UPPER({make_col}{r})=UPPER({make_ext_col}{r})"
                    )

            if "Model Match" in headers:
                model_match_col = get_column_letter(headers["Model Match"])
                for r in range(2, max_row + 1):
                    ws[f"{model_match_col}{r}"] = (
                        f"=UPPER({model_col}{r})=UPPER({model_ext_col}{r})"
                    )

            if "Year Match" in headers:
                year_match_col = get_column_letter(headers["Year Match"])
                for r in range(2, max_row + 1):
                    ws[f"{year_match_col}{r}"] = (
                        f'=TEXT({year_col}{r},"0")=TEXT({year_ext_col}{r},"0")'
                    )

            # Add formulas for extra field match columns
            if extra_match_cols:
                for ref_col, logs_col in extra_match_cols:
                    match_name = f"{logs_col}_match"
                    if match_name in headers:
                        match_col_letter = get_column_letter(headers[match_name])
                        # Handle suffixed column names from merge (when both sides share the same name)
                        logs_header = logs_col + "_primary" if logs_col + "_primary" in headers else logs_col
                        ref_header = ref_col + "_join" if ref_col + "_join" in headers else ref_col
                        if logs_header in headers and ref_header in headers:
                            logs_col_letter = get_column_letter(headers[logs_header])
                            ref_col_letter = get_column_letter(headers[ref_header])
                            for r in range(2, max_row + 1):
                                ws[f"{match_col_letter}{r}"] = (
                                    f"=UPPER({logs_col_letter}{r})=UPPER({ref_col_letter}{r})"
                                )

        # Create Summary sheet
        summary_metrics = [
            "Total Records",
            "Make Matches",
            "Model Matches",
            "Year Matches",
        ]
        summary_counts = [
            len(self.merged_data),
            make_matches_calc.sum(),
            model_matches_calc.sum(),
            year_matches_calc.sum(),
        ]
        summary_percentages = [100.0]

        # Add extra field match counts and percentages to summary
        extra_match_names = []
        for ref_col, logs_col in extra_match_cols:
            match_name = f"{logs_col}_match"
            extra_match_names.append(match_name)
            summary_metrics.append(f"{logs_col} Matches")
            summary_counts.append(
                extra_matches[match_name].sum() if match_name in extra_matches else 0
            )

        # Calculate percentages for Make, Model, Year
        total = len(self.merged_data)
        summary_percentages.append(
            make_matches_calc.sum() / total * 100 if total > 0 else 0
        )
        summary_percentages.append(
            model_matches_calc.sum() / total * 100 if total > 0 else 0
        )
        summary_percentages.append(
            year_matches_calc.sum() / total * 100 if total > 0 else 0
        )

        # Calculate percentages for extra fields
        for match_name in extra_match_names:
            summary_percentages.append(
                extra_matches[match_name].sum() / total * 100 if total > 0 else 0
            )

        summary_data = {
            "Metric": summary_metrics,
            "Count": summary_counts,
            "Percentage": summary_percentages,
        }
        summary_df = pd.DataFrame(summary_data)

        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        for r_idx, row in enumerate(summary_df.values, 1):
            for c_idx, value in enumerate(row, 1):
                summary_ws.cell(row=r_idx, column=c_idx, value=value)

        # Add headers to summary
        for c_idx, header in enumerate(summary_df.columns, 1):
            summary_ws.cell(row=1, column=c_idx, value=header)

        # Apply dynamic conditional formatting to Analysis sheet
        self._apply_dynamic_conditional_formatting(wb, ws, include_mask_in_main)

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _apply_dynamic_conditional_formatting(
        self, workbook, worksheet, include_mask_in_main: bool = True
    ) -> None:
        """Apply dynamic conditional formatting using Excel formulas that compare columns directly"""
        max_row = worksheet.max_row

        # Map headers to column letters
        headers = {cell.value: cell.column for cell in worksheet[1]}

        # Get extra_match_cols
        extra_match_cols = getattr(self, "extra_match_cols", [])

        # Define colors
        mint_green = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        coral_pink = PatternFill(
            start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"
        )

        # Get column letters for data columns
        make_col = get_column_letter(headers.get("Make", 1))
        make_ext_col = get_column_letter(headers.get("Make_ext", 1))
        model_col = get_column_letter(headers.get("Model", 1))
        model_ext_col = get_column_letter(headers.get("Model_ext", 1))
        year_col = get_column_letter(headers.get("ModelYear", 1))
        year_ext_col = get_column_letter(headers.get("ModelYear_ext", 1))

        if include_mask_in_main:
            # Use the match columns directly from the main sheet with dynamic formulas as backup
            format_mappings = [
                (
                    "Make_ext",
                    "Make Match",
                    make_col,
                    make_ext_col,
                ),  # Make_ext column uses Make Match column
                (
                    "Model_ext",
                    "Model Match",
                    model_col,
                    model_ext_col,
                ),  # Model_ext column uses Model Match column
                (
                    "ModelYear_ext",
                    "Year Match",
                    year_col,
                    year_ext_col,
                ),  # ModelYear_ext column uses Year Match column
            ]

            # Add extra field mappings (handle suffixed column names from merge)
            # Formatting goes on the reference (_join) side, matching Make_ext/Model_ext pattern
            for ref_col, logs_col in extra_match_cols:
                match_name = f"{logs_col}_match"
                logs_header = logs_col + "_primary" if logs_col + "_primary" in headers else logs_col
                ref_header = ref_col + "_join" if ref_col + "_join" in headers else ref_col
                if logs_header in headers and ref_header in headers:
                    format_mappings.append(
                        (
                            ref_header,
                            match_name,
                            get_column_letter(headers[logs_header]),
                            get_column_letter(headers[ref_header]),
                        )
                    )

            for target_col, match_col, source_col, target_col_letter in format_mappings:
                if target_col in headers:
                    target_letter = get_column_letter(headers[target_col])
                    cell_range = f"{target_letter}2:{target_letter}{max_row}"

                    if match_col in headers:
                        # Use the boolean match columns if they exist
                        match_letter = get_column_letter(headers[match_col])
                        green_rule = FormulaRule(
                            formula=[f"=${match_letter}2=TRUE"], fill=mint_green
                        )
                        pink_rule = FormulaRule(
                            formula=[f"=${match_letter}2=FALSE"], fill=coral_pink
                        )
                    else:
                        # Fallback to dynamic formula comparison if match columns don't exist
                        if target_col == "ModelYear_ext":
                            # Year comparison as text
                            green_formula = (
                                f'=TEXT({source_col}2,"0")=TEXT({target_letter}2,"0")'
                            )
                            pink_formula = (
                                f'=TEXT({source_col}2,"0")<>TEXT({target_letter}2,"0")'
                            )
                        else:
                            # Make/Model comparison case-insensitive
                            green_formula = (
                                f"=UPPER({source_col}2)=UPPER({target_letter}2)"
                            )
                            pink_formula = (
                                f"=UPPER({source_col}2)<>UPPER({target_letter}2)"
                            )

                        green_rule = FormulaRule(
                            formula=[green_formula], fill=mint_green
                        )
                        pink_rule = FormulaRule(formula=[pink_formula], fill=coral_pink)

                    worksheet.conditional_formatting.add(cell_range, green_rule)
                    worksheet.conditional_formatting.add(cell_range, pink_rule)
        else:
            # Create separate Mask sheet for dynamic formulas (original behavior)
            mask_ws = workbook.create_sheet("Mask")
            mask_ws["A1"] = "MakeMatch"
            mask_ws["B1"] = "ModelMatch"
            mask_ws["C1"] = "YearMatch"

            # Add extra match columns to mask sheet
            mask_col_idx = 4  # Start after A, B, C
            extra_match_cols_list = []
            for ref_col, logs_col in extra_match_cols:
                col_letter = get_column_letter(mask_col_idx)
                mask_ws[f"{col_letter}1"] = f"{logs_col}Match"
                extra_match_cols_list.append((logs_col, ref_col, col_letter))
                mask_col_idx += 1

            # Build mask formulas - directly compare the data columns
            for r in range(2, max_row + 1):
                # Make match formula - compare Make vs Make_ext (case-insensitive)
                mask_ws[f"A{r}"] = (
                    f"=UPPER(Analysis!{make_col}{r})=UPPER(Analysis!{make_ext_col}{r})"
                )
                # Model match formula - compare Model vs Make_ext (case-insensitive)
                mask_ws[f"B{r}"] = (
                    f"=UPPER(Analysis!{model_col}{r})=UPPER(Analysis!{model_ext_col}{r})"
                )
                # Year match formula - compare ModelYear vs ModelYear_ext (as text)
                mask_ws[f"C{r}"] = (
                    f'=TEXT(Analysis!{year_col}{r},"0")=TEXT(Analysis!{year_ext_col}{r},"0")'
                )

                # Add extra field match formulas
                for logs_col, ref_col, mask_col_letter in extra_match_cols_list:
                    # Handle suffixed column names from merge
                    logs_header = logs_col + "_primary" if logs_col + "_primary" in headers else logs_col
                    ref_header = ref_col + "_join" if ref_col + "_join" in headers else ref_col
                    if logs_header in headers and ref_header in headers:
                        logs_col_letter = get_column_letter(headers[logs_header])
                        ref_col_letter = get_column_letter(headers[ref_header])
                        mask_ws[f"{mask_col_letter}{r}"] = (
                            f"=UPPER(Analysis!{logs_col_letter}{r})=UPPER(Analysis!{ref_col_letter}{r})"
                        )

            # Apply conditional formatting to specific columns using dynamic formulas
            format_mappings = [
                (
                    "Make_ext",
                    make_col,
                    make_ext_col,
                ),  # Make_ext column compared to Make column
                (
                    "Model_ext",
                    model_col,
                    model_ext_col,
                ),  # Model_ext column compared to Model column
                (
                    "ModelYear_ext",
                    year_col,
                    year_ext_col,
                ),  # ModelYear_ext column compared to ModelYear column
            ]

            # Add extra field mappings (handle suffixed column names from merge)
            # Formatting goes on the reference (_join) side, matching Make_ext/Model_ext pattern
            for logs_col, ref_col, _ in extra_match_cols_list:
                logs_header = logs_col + "_primary" if logs_col + "_primary" in headers else logs_col
                ref_header = ref_col + "_join" if ref_col + "_join" in headers else ref_col
                if logs_header in headers and ref_header in headers:
                    format_mappings.append(
                        (
                            ref_header,
                            get_column_letter(headers[logs_header]),
                            get_column_letter(headers[ref_header]),
                        )
                    )

            for target_col, source_col, target_col_letter in format_mappings:
                if target_col in headers:
                    target_letter = get_column_letter(headers[target_col])
                    cell_range = f"{target_letter}2:{target_letter}{max_row}"

                    # Green for matches - dynamic formula comparing columns
                    if target_col == "ModelYear_ext":
                        # Year comparison as text
                        green_formula = (
                            f'=TEXT({source_col}2,"0")=TEXT({target_letter}2,"0")'
                        )
                        pink_formula = (
                            f'=TEXT({source_col}2,"0")<>TEXT({target_letter}2,"0")'
                        )
                    else:
                        # Make/Model comparison case-insensitive
                        green_formula = f"=UPPER({source_col}2)=UPPER({target_letter}2)"
                        pink_formula = f"=UPPER({source_col}2)<>UPPER({target_letter}2)"

                    green_rule = FormulaRule(formula=[green_formula], fill=mint_green)
                    worksheet.conditional_formatting.add(cell_range, green_rule)

                    pink_rule = FormulaRule(formula=[pink_formula], fill=coral_pink)
                    worksheet.conditional_formatting.add(cell_range, pink_rule)

    def get_verification_summary(self) -> Dict[str, Union[int, float]]:
        """Get summary statistics from verification"""
        if self.merged_data is None:
            return {}

        # Get extra_match_cols
        extra_match_cols = getattr(self, "extra_match_cols", [])

        # Calculate matches dynamically
        make_matches_calc = (
            self.merged_data["Make"].str.lower()
            == self.merged_data["Make_ext"].str.lower()
        )
        model_matches_calc = (
            self.merged_data["Model"].str.lower()
            == self.merged_data["Model_ext"].str.lower()
        )
        year_matches_calc = self.merged_data["ModelYear"].astype(
            str
        ) == self.merged_data["ModelYear_ext"].astype(str)

        # Calculate extra field matches
        extra_matches = {}
        extra_match_cols = getattr(self, "extra_match_cols", [])
        if extra_match_cols:
            merged_cols = list(self.merged_data.columns)
            for ref_col, logs_col in extra_match_cols:
                match_name = f"{logs_col}_match"
                ref_col_actual = self._resolve_merged_column(
                    ref_col, merged_cols, "_join"
                )
                logs_col_actual = self._resolve_merged_column(
                    logs_col, merged_cols, "_primary"
                )
                ref_vals = self.merged_data[ref_col_actual].astype(str).str.lower()
                logs_vals = self.merged_data[logs_col_actual].astype(str).str.lower()
                extra_matches[match_name] = ref_vals == logs_vals

        total_records = len(self.merged_data)
        make_matches = make_matches_calc.sum()
        model_matches = model_matches_calc.sum()
        year_matches = year_matches_calc.sum()

        result = {
            "total_records": total_records,
            "make_matches": make_matches,
            "model_matches": model_matches,
            "year_matches": year_matches,
            "make_match_percentage": make_matches / total_records * 100
            if total_records > 0
            else 0,
            "model_match_percentage": model_matches / total_records * 100
            if total_records > 0
            else 0,
            "year_match_percentage": year_matches / total_records * 100
            if total_records > 0
            else 0,
        }

        # Add extra field match info
        for match_name, match_series in extra_matches.items():
            result[match_name] = match_series.sum()
            result[f"{match_name}_percentage"] = (
                match_series.sum() / total_records * 100 if total_records > 0 else 0
            )

        return result

    def get_sample_mismatches(self, n: int = 5) -> pd.DataFrame:
        """Get sample mismatches for display"""
        if self.merged_data is None:
            return pd.DataFrame()

        # Get extra_match_cols
        extra_match_cols = getattr(self, "extra_match_cols", [])

        # Calculate mismatches dynamically
        make_matches = (
            self.merged_data["Make"].str.lower()
            == self.merged_data["Make_ext"].str.lower()
        )
        model_matches = (
            self.merged_data["Model"].str.lower()
            == self.merged_data["Model_ext"].str.lower()
        )
        year_matches = self.merged_data["ModelYear"].astype(str) == self.merged_data[
            "ModelYear_ext"
        ].astype(str)

        # Add extra field mismatches
        mismatch_condition = ~make_matches | ~model_matches | ~year_matches
        if extra_match_cols:
            merged_cols = list(self.merged_data.columns)
            for ref_col, logs_col in extra_match_cols:
                ref_col_actual = self._resolve_merged_column(
                    ref_col, merged_cols, "_join"
                )
                logs_col_actual = self._resolve_merged_column(
                    logs_col, merged_cols, "_primary"
                )
                ref_vals = self.merged_data[ref_col_actual].astype(str).str.lower()
                logs_vals = self.merged_data[logs_col_actual].astype(str).str.lower()
                mismatch_condition = mismatch_condition | (ref_vals != logs_vals)

        mismatches = self.merged_data[mismatch_condition]

        if len(mismatches) == 0:
            return pd.DataFrame()

        sample_cols = [
            "VIN",
            "Make",
            "Make_ext",
            "Model",
            "Model_ext",
            "ModelYear",
            "ModelYear_ext",
        ]

        # Add extra field columns to sample
        for ref_col, logs_col in extra_match_cols:
            if logs_col in self.merged_data.columns:
                sample_cols.append(logs_col)
            if ref_col in self.merged_data.columns:
                sample_cols.append(ref_col)

        # Get available columns
        available_cols = [col for col in sample_cols if col in self.merged_data.columns]

        return mismatches[available_cols].head(n)
